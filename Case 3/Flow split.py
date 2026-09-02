import pandas as pd
import zipfile
import io
import openpyxl


# ============================================================
# ALL-KANNS LIFE INSURANCE
# MIXED LUMP-SUM / PENSION LIABILITY SCENARIO MODEL
# ============================================================


# ============================================================
# 1. FILE LOCATIONS
# ============================================================

MORTALITY_FILE = "germany_mortality.csv"

EIOPA_ZIP_FILE = "EIOPA_RFR_20260731.zip"

EIOPA_EXCEL_FILE = "EIOPA_RFR_20260731_Term_Structures.xlsx"

EIOPA_SHEET = "RFR_spot_no_VA"


# ============================================================
# 2. CASE INPUTS
# ============================================================

N_MALE = 50_000
N_FEMALE = 50_000

TOTAL_POLICYHOLDERS = (
    N_MALE + N_FEMALE
)

CURRENT_AGE = 50
RETIREMENT_AGE = 65

YEARS_TO_RETIREMENT = (
    RETIREMENT_AGE - CURRENT_AGE
)

MAX_AGE = 100

INITIAL_CAPITAL_PER_PERSON = 50_000

ANNUAL_CONTRIBUTION = 5_000

CONTRIBUTION_YEARS = 10

# Professor clarification:
# minimum 1% return on contributed capital
GUARANTEED_ACCUMULATION_RATE = 0.01

# Pension paid for at least 10 years
PENSION_GUARANTEE_YEARS = 10

# ------------------------------------------------------------
# Pension conversion assumption
#
# This determines the annual pension amount.
#
# We currently use 1%.
# ------------------------------------------------------------

PENSION_CONVERSION_RATE = 0.01


# ============================================================
# 3. SCENARIOS
# ============================================================

# Percentage choosing lump sum

LUMP_SUM_SCENARIOS = [

    0.00,
    0.25,
    0.50,
    0.55,
    0.75,
    1.00

]


# ============================================================
# 4. GUARANTEED CAPITAL AT AGE 65
# ============================================================

# Existing EUR 50,000 grows for 15 years

FV_INITIAL_PER_PERSON = (

    INITIAL_CAPITAL_PER_PERSON

    *

    (
        1
        +
        GUARANTEED_ACCUMULATION_RATE
    )
    ** YEARS_TO_RETIREMENT

)


# ------------------------------------------------------------
# Future contributions
#
# Contribution 1 grows for 14 years
# ...
# Contribution 10 grows for 5 years
# ------------------------------------------------------------

FV_CONTRIBUTIONS_PER_PERSON = 0

contribution_details = []


for contribution_year in range(
    1,
    CONTRIBUTION_YEARS + 1
):

    years_of_growth = (

        YEARS_TO_RETIREMENT
        -
        contribution_year

    )

    future_value = (

        ANNUAL_CONTRIBUTION

        *

        (
            1
            +
            GUARANTEED_ACCUMULATION_RATE
        )
        ** years_of_growth

    )

    FV_CONTRIBUTIONS_PER_PERSON += (
        future_value
    )


    contribution_details.append({

        "Contribution_Year":
            contribution_year,

        "Contribution":
            ANNUAL_CONTRIBUTION,

        "Years_of_Growth":
            years_of_growth,

        "Value_at_Age_65":
            future_value

    })


# ------------------------------------------------------------
# Guaranteed capital per person at 65
# ------------------------------------------------------------

GUARANTEED_CAPITAL_PER_PERSON = (

    FV_INITIAL_PER_PERSON

    +

    FV_CONTRIBUTIONS_PER_PERSON

)


# ------------------------------------------------------------
# Total guaranteed capital
# ------------------------------------------------------------

TOTAL_GUARANTEED_CAPITAL_AT_65 = (

    GUARANTEED_CAPITAL_PER_PERSON

    *

    TOTAL_POLICYHOLDERS

)


# ============================================================
# 5. LOAD MORTALITY DATA
# ============================================================

mortality_input = pd.read_csv(
    MORTALITY_FILE
)


mortality_input = mortality_input[

    (
        mortality_input["Age"]
        >= RETIREMENT_AGE
    )

    &

    (
        mortality_input["Age"]
        < MAX_AGE
    )

].copy()


# ------------------------------------------------------------
# Check mortality ages
# ------------------------------------------------------------

required_ages = set(

    range(
        RETIREMENT_AGE,
        MAX_AGE
    )

)

available_ages = set(
    mortality_input["Age"]
)

missing_ages = (

    required_ages
    -
    available_ages

)


if missing_ages:

    raise ValueError(

        f"Missing mortality rates for ages: "
        f"{sorted(missing_ages)}"

    )


# ============================================================
# 6. MORTALITY DICTIONARIES
# ============================================================

male_qx = dict(

    zip(

        mortality_input["Age"],

        mortality_input["Male_qx"]

    )

)


female_qx = dict(

    zip(

        mortality_input["Age"],

        mortality_input["Female_qx"]

    )

)


# ============================================================
# 7. SURVIVAL FUNCTION
# ============================================================

def calculate_survival(
    start_population,
    qx_by_age
):

    alive = float(
        start_population
    )

    rows = []


    for age in range(
        RETIREMENT_AGE,
        MAX_AGE
    ):

        qx = qx_by_age[age]

        expected_deaths = (
            alive * qx
        )

        alive_next = (
            alive
            -
            expected_deaths
        )


        rows.append({

            "Age":
                age,

            "Alive_Start":
                alive,

            "qx":
                qx,

            "Expected_Deaths":
                expected_deaths,

            "Alive_Next":
                alive_next

        })


        alive = alive_next


    return pd.DataFrame(
        rows
    )


# ============================================================
# 8. BASE SURVIVAL CURVES FOR 100% PENSION POPULATION
# ============================================================

male_base = calculate_survival(
    N_MALE,
    male_qx
)

female_base = calculate_survival(
    N_FEMALE,
    female_qx
)


# ============================================================
# 9. READ EIOPA CURVE
# ============================================================

with zipfile.ZipFile(
    EIOPA_ZIP_FILE,
    "r"
) as z:

    excel_bytes = z.read(
        EIOPA_EXCEL_FILE
    )


workbook = openpyxl.load_workbook(

    io.BytesIO(
        excel_bytes
    ),

    data_only=True

)


sheet = workbook[
    EIOPA_SHEET
]


# ============================================================
# 10. EXTRACT EUR EIOPA SPOT CURVE
# ============================================================

eiopa_curve = {}

row = 11


while True:

    maturity = sheet.cell(
        row=row,
        column=2
    ).value

    euro_rate = sheet.cell(
        row=row,
        column=3
    ).value


    if maturity is None:

        break


    if euro_rate is not None:

        eiopa_curve[
            int(maturity)
        ] = float(
            euro_rate
        )


    row += 1


# ============================================================
# 11. EIOPA INFORMATION
# ============================================================

EIOPA_LLP = sheet["C5"].value

EIOPA_CONVERGENCE = sheet["C6"].value

EIOPA_UFR = sheet["C7"].value


# ============================================================
# 12. EIOPA HELPER FUNCTIONS
# ============================================================

def get_eiopa_rate(year):

    year = int(year)

    if year not in eiopa_curve:

        raise ValueError(

            f"EIOPA curve does not contain "
            f"maturity {year}."

        )

    return eiopa_curve[year]


def get_discount_factor(year):

    year = int(year)

    rate = get_eiopa_rate(
        year
    )

    return (

        1

        /

        (
            1 + rate
        )
        ** year

    )


# ============================================================
# 13. YEAR-15 EIOPA DISCOUNT FACTOR
# ============================================================
#
# This is the discount factor from TODAY to Year 15.
# ============================================================

DF_0_15 = get_discount_factor(
    YEARS_TO_RETIREMENT
)


EIOPA_RATE_15 = get_eiopa_rate(
    YEARS_TO_RETIREMENT
)


# ============================================================
# 14. CALCULATE PENSION PER PERSON
# ============================================================
#
# IMPORTANT:
#
# The pension per pension-taking policyholder should NOT
# depend on how many other people choose the lump sum.
#
# Therefore we first calculate the pension using the
# full 100,000-person population.
#
# The resulting pension per person is then kept fixed
# across all scenarios.
# ============================================================

pension_base = pd.DataFrame()


pension_base["Age"] = (
    male_base["Age"]
)


pension_base["Pension_Year"] = (

    pension_base["Age"]

    -

    RETIREMENT_AGE

    +

    1

)


pension_base["Male_Alive"] = (
    male_base["Alive_Start"]
)

pension_base["Female_Alive"] = (
    female_base["Alive_Start"]
)


pension_base["Total_Alive"] = (

    pension_base["Male_Alive"]

    +

    pension_base["Female_Alive"]

)


# ------------------------------------------------------------
# Guaranteed pension units
# ------------------------------------------------------------

def base_pension_units(row):

    if (
        row["Pension_Year"]
        <= PENSION_GUARANTEE_YEARS
    ):

        return (
            TOTAL_POLICYHOLDERS
        )

    return (
        row["Total_Alive"]
    )


pension_base["Pension_Units"] = (

    pension_base.apply(
        base_pension_units,
        axis=1
    )

)


# ------------------------------------------------------------
# Conversion discount factors
# ------------------------------------------------------------

pension_base[
    "Conversion_DF"
] = (

    1

    /

    (
        1
        +
        PENSION_CONVERSION_RATE
    )
    **
    pension_base["Pension_Year"]

)


pension_base[
    "PV_1EUR"
] = (

    pension_base["Pension_Units"]

    *

    pension_base["Conversion_DF"]

)


POOL_ANNUITY_FACTOR = (

    pension_base[
        "PV_1EUR"
    ].sum()

)


ANNUAL_PENSION_PER_PERSON = (

    TOTAL_GUARANTEED_CAPITAL_AT_65

    /

    POOL_ANNUITY_FACTOR

)


MONTHLY_PENSION_PER_PERSON = (

    ANNUAL_PENSION_PER_PERSON

    /

    12

)


# ============================================================
# 15. SCENARIO ENGINE
# ============================================================

scenario_results = []

scenario_cashflows = {}


for lump_sum_share in LUMP_SUM_SCENARIOS:

    # ========================================================
    # 15A. SPLIT POPULATION
    # ========================================================

    pension_share = (
        1
        -
        lump_sum_share
    )


    number_lump_sum = (

        TOTAL_POLICYHOLDERS

        *

        lump_sum_share

    )


    number_pension = (

        TOTAL_POLICYHOLDERS

        *

        pension_share

    )


    # ========================================================
    # 15B. MALE/FEMALE PENSION POPULATION
    # ========================================================

    male_pensioners = (

        N_MALE

        *

        pension_share

    )


    female_pensioners = (

        N_FEMALE

        *

        pension_share

    )


    # ========================================================
    # 15C. LUMP-SUM PAYMENT AT YEAR 15
    # ========================================================

    lump_sum_at_15 = (

        number_lump_sum

        *

        GUARANTEED_CAPITAL_PER_PERSON

    )


    # ========================================================
    # 15D. PV TODAY OF LUMP-SUM LIABILITY
    # ========================================================

    pv_lump_sum_today = (

        lump_sum_at_15

        *

        DF_0_15

    )


    # ========================================================
    # 15E. BUILD PENSION SURVIVAL MODEL
    # ========================================================

    male_scenario = calculate_survival(

        male_pensioners,

        male_qx

    )


    female_scenario = calculate_survival(

        female_pensioners,

        female_qx

    )


    scenario = pd.DataFrame()


    scenario["Age"] = (
        male_scenario["Age"]
    )


    scenario["Pension_Year"] = (

        scenario["Age"]

        -

        RETIREMENT_AGE

        +

        1

    )


    # --------------------------------------------------------
    # Years from today
    #
    # Pension Year 1 occurs at end of first pension year.
    #
    # Therefore:
    #
    # Pension Year 1 = Year 16 from today.
    #
    # --------------------------------------------------------

    scenario[
        "Years_From_Today"
    ] = (

        YEARS_TO_RETIREMENT

        +

        scenario["Pension_Year"]

    )


    scenario["Male_Alive"] = (
        male_scenario["Alive_Start"]
    )

    scenario["Female_Alive"] = (
        female_scenario["Alive_Start"]
    )


    scenario["Total_Alive"] = (

        scenario["Male_Alive"]

        +

        scenario["Female_Alive"]

    )


    scenario["Male_Deaths"] = (
        male_scenario["Expected_Deaths"]
    )

    scenario["Female_Deaths"] = (
        female_scenario["Expected_Deaths"]
    )


    scenario["Total_Deaths"] = (

        scenario["Male_Deaths"]

        +

        scenario["Female_Deaths"]

    )


    # ========================================================
    # 15F. NUMBER OF PENSION PAYMENTS
    # ========================================================

    def scenario_pension_units(row):

        if (
            row["Pension_Year"]
            <=
            PENSION_GUARANTEE_YEARS
        ):

            return (
                number_pension
            )

        return (
            row["Total_Alive"]
        )


    scenario["Pension_Units"] = (

        scenario.apply(
            scenario_pension_units,
            axis=1
        )

    )


    # ========================================================
    # 15G. PENSION CASH FLOWS
    # ========================================================

    scenario[
        "Pension_Cashflow"
    ] = (

        scenario["Pension_Units"]

        *

        ANNUAL_PENSION_PER_PERSON

    )


    # ========================================================
    # 15H. EIOPA SPOT RATE
    # ========================================================

    scenario[
        "EIOPA_Spot_Rate"
    ] = (

        scenario[
            "Years_From_Today"
        ].apply(
            get_eiopa_rate
        )

    )


    # ========================================================
    # 15I. DISCOUNT FACTOR FROM TODAY
    # ========================================================

    scenario[
        "DF_0_t"
    ] = (

        1

        /

        (
            1
            +
            scenario[
                "EIOPA_Spot_Rate"
            ]
        )

        **

        scenario[
            "Years_From_Today"
        ]

    )


    # ========================================================
    # 15J. PV TODAY OF PENSION CASH FLOWS
    # ========================================================

    scenario[
        "PV_Pension_Today"
    ] = (

        scenario[
            "Pension_Cashflow"
        ]

        *

        scenario[
            "DF_0_t"
        ]

    )


    pv_pension_today = (

        scenario[
            "PV_Pension_Today"
        ].sum()

    )


    # ========================================================
    # 15K. FORWARD DISCOUNT FACTOR FROM YEAR 15
    # ========================================================
    #
    # Forward discount factor:
    #
    # DF(15,t)
    #
    # =
    #
    # DF(0,t) / DF(0,15)
    #
    # This values future pension cash flows
    # as of Year 15 using today's EIOPA curve.
    # ========================================================

    scenario[
        "DF_15_t"
    ] = (

        scenario[
            "DF_0_t"
        ]

        /

        DF_0_15

    )


    # ========================================================
    # 15L. PV AT YEAR 15 OF PENSION LIABILITY
    # ========================================================

    scenario[
        "PV_Pension_Year15"
    ] = (

        scenario[
            "Pension_Cashflow"
        ]

        *

        scenario[
            "DF_15_t"
        ]

    )


    pv_pension_year15 = (

        scenario[
            "PV_Pension_Year15"
        ].sum()

    )


    # ========================================================
    # 15M. TOTAL LIABILITY AT YEAR 15
    # ========================================================
    #
    # Lump sum is physically paid at Year 15.
    #
    # Pension liability is the Year-15 PV
    # of all remaining pension cash flows.
    # ========================================================

    total_liability_year15 = (

        lump_sum_at_15

        +

        pv_pension_year15

    )


    # ========================================================
    # 15N. TOTAL PV TODAY
    # ========================================================

    total_pv_today = (

        pv_lump_sum_today

        +

        pv_pension_today

    )


    # ========================================================
    # 15O. NOMINAL PENSION PAYMENTS
    # ========================================================

    nominal_pension_payments = (

        scenario[
            "Pension_Cashflow"
        ].sum()

    )


    # ========================================================
    # 15P. TOTAL EXPECTED BENEFIT PAYMENTS
    # ========================================================
    #
    # This combines:
    #
    # Lump sum at 65
    #
    # +
    #
    # all expected nominal pension payments.
    # ========================================================

    total_nominal_benefits = (

        lump_sum_at_15

        +

        nominal_pension_payments

    )


    # ========================================================
    # 15Q. STORE RESULTS
    # ========================================================

    scenario_results.append({

        "Lump_Sum_%":
            lump_sum_share * 100,

        "Pension_%":
            pension_share * 100,

        "Lump_Sum_People":
            number_lump_sum,

        "Pension_People":
            number_pension,

        "Lump_Sum_at_Year15":
            lump_sum_at_15,

        "PV_Lump_Sum_Today":
            pv_lump_sum_today,

        "PV_Pension_Year15":
            pv_pension_year15,

        "Total_Liability_Year15":
            total_liability_year15,

        "PV_Pension_Today":
            pv_pension_today,

        "Total_PV_Today":
            total_pv_today,

        "Nominal_Pension_Payments":
            nominal_pension_payments,

        "Total_Nominal_Benefits":
            total_nominal_benefits

    })


    # --------------------------------------------------------
    # Store detailed cash flows
    # --------------------------------------------------------

    scenario_cashflows[
        lump_sum_share
    ] = scenario


# ============================================================
# 16. CREATE SUMMARY DATAFRAME
# ============================================================

summary = pd.DataFrame(
    scenario_results
)


# ============================================================
# 17. PRINT GENERAL RESULTS
# ============================================================

print(
    "\n"
    +
    "=" * 90
)

print(
    "ALL-KANNS LIFE INSURANCE"
)

print(
    "MIXED LUMP-SUM / PENSION LIABILITY MODEL"
)

print(
    "=" * 90
)


print(
    f"\nGuaranteed accumulation rate: "
    f"{GUARANTEED_ACCUMULATION_RATE * 100:.2f}%"
)


print(
    f"Guaranteed capital per person at age 65: "
    f"EUR {GUARANTEED_CAPITAL_PER_PERSON:,.2f}"
)


print(
    f"Total guaranteed capital at age 65: "
    f"EUR {TOTAL_GUARANTEED_CAPITAL_AT_65:,.2f}"
)


print(
    f"\nAnnual pension per pensioner: "
    f"EUR {ANNUAL_PENSION_PER_PERSON:,.2f}"
)


print(
    f"Monthly pension per pensioner: "
    f"EUR {MONTHLY_PENSION_PER_PERSON:,.2f}"
)


print(
    f"\nEIOPA 15Y spot rate: "
    f"{EIOPA_RATE_15 * 100:.4f}%"
)


print(
    f"EIOPA discount factor today -> Year 15: "
    f"{DF_0_15:.6f}"
)


print(
    f"EIOPA LLP: "
    f"{EIOPA_LLP} years"
)


print(
    f"EIOPA UFR: "
    f"{EIOPA_UFR:.2f}%"
)


# ============================================================
# 18. PRINT SCENARIO SUMMARY IN BILLIONS
# ============================================================

summary_billions = summary.copy()


money_columns = [

    "Lump_Sum_at_Year15",

    "PV_Lump_Sum_Today",

    "PV_Pension_Year15",

    "Total_Liability_Year15",

    "PV_Pension_Today",

    "Total_PV_Today",

    "Nominal_Pension_Payments",

    "Total_Nominal_Benefits"

]


for column in money_columns:

    summary_billions[column] = (

        summary_billions[column]

        /

        1_000_000_000

    )


print(
    "\n"
    +
    "=" * 90
)

print(
    "SCENARIO SUMMARY - EUR BILLIONS"
)

print(
    "=" * 90
    +
    "\n"
)


print(

    summary_billions

    .round(4)

    .to_string(
        index=False
    )

)


# ============================================================
# 19. PRINT A CLEANER PRESENTATION TABLE
# ============================================================

presentation_table = summary_billions[[

    "Lump_Sum_%",

    "Pension_%",

    "Lump_Sum_at_Year15",

    "PV_Pension_Year15",

    "Total_Liability_Year15",

    "PV_Lump_Sum_Today",

    "PV_Pension_Today",

    "Total_PV_Today"

]].copy()


presentation_table.columns = [

    "Lump Sum %",

    "Pension %",

    "Lump Sum at Y15 (€bn)",

    "PV Pension at Y15 (€bn)",

    "Total Liability at Y15 (€bn)",

    "PV Lump Sum Today (€bn)",

    "PV Pension Today (€bn)",

    "Total PV Today (€bn)"

]


print(
    "\n"
    +
    "=" * 90
)

print(
    "PRESENTATION TABLE"
)

print(
    "=" * 90
    +
    "\n"
)


print(

    presentation_table

    .round(3)

    .to_string(
        index=False
    )

)


# ============================================================
# 20. EXPORT SUMMARY TO EXCEL
# ============================================================

with pd.ExcelWriter(
    "mixed_liability_scenarios.xlsx",
    engine="openpyxl"
) as writer:

    # Full summary

    summary.to_excel(

        writer,

        sheet_name="Scenario Summary",

        index=False

    )


    # Summary in billions

    presentation_table.to_excel(

        writer,

        sheet_name="Presentation Table",

        index=False

    )


    # Contribution calculation

    pd.DataFrame(
        contribution_details
    ).to_excel(

        writer,

        sheet_name="Accumulation",

        index=False

    )


    # EIOPA curve

    pd.DataFrame({

        "Maturity":
            list(
                eiopa_curve.keys()
            ),

        "EUR_Spot_Rate":
            list(
                eiopa_curve.values()
            )

    }).to_excel(

        writer,

        sheet_name="EIOPA Curve",

        index=False

    )


    # --------------------------------------------------------
    # Detailed scenario cash flows
    # --------------------------------------------------------

    for lump_sum_share, df in scenario_cashflows.items():

        sheet_name = (

            f"LS_{int(lump_sum_share * 100)}pct"

        )

        df.to_excel(

            writer,

            sheet_name=sheet_name,

            index=False

        )


# ============================================================
# 21. FINISH
# ============================================================

print(
    "\n"
    +
    "=" * 90
)

print(
    "MODEL COMPLETE"
)

print(
    "=" * 90
)


print(
    "\nOutput file:"
)

print(
    "mixed_liability_scenarios.xlsx"
)