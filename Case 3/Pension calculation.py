import pandas as pd
import zipfile
import io
import openpyxl

# ============================================================
# ALL-KANNS LIFE INSURANCE
# COMPLETE PENSION LIABILITY MODEL
# ============================================================


# ============================================================
# 1. FILE LOCATIONS
# ============================================================

MORTALITY_FILE = "germany_mortality.csv"

EIOPA_ZIP_FILE = "EIOPA_RFR_20260731.zip"

EIOPA_EXCEL_FILE = "EIOPA_RFR_20260731_Term_Structures.xlsx"

EIOPA_SHEET = "RFR_spot_no_VA"


# ============================================================
# 2. CASE INPUTS
# ============================================================

N_MALE = 50_000
N_FEMALE = 50_000
TOTAL_POLICYHOLDERS = N_MALE + N_FEMALE

CURRENT_AGE = 50
RETIREMENT_AGE = 65

# Mortality model currently runs until age 100
MAX_AGE = 100

INITIAL_CAPITAL_PER_PERSON = 50_000

ANNUAL_CONTRIBUTION = 5_000

CONTRIBUTION_YEARS = 10

# Professor's clarification:
# minimum 1% return on contributed capital
GUARANTEED_ACCUMULATION_RATE = 0.01

# Pension is paid for at least 10 years
PENSION_GUARANTEE_YEARS = 10


# ============================================================
# 3. PENSION CONVERSION ASSUMPTION
# ============================================================
#
# IMPORTANT:
#
# The 1% guaranteed accumulation tells us how much guaranteed
# capital the policyholder has accumulated by age 65.
#
# The case does not explicitly provide a pension conversion
# rate.
#
# For now we use 1% as the actuarial conversion assumption.
#
# THIS RATE DETERMINES THE PENSION AMOUNT.
#
# It does NOT determine the market PV of the liability.
#
# Market PV is calculated separately using the EIOPA curve.
# ============================================================

PENSION_CONVERSION_RATE = 0.01


# ============================================================
# 4. CALCULATE GUARANTEED CAPITAL AT AGE 65
# ============================================================

YEARS_TO_RETIREMENT = (
    RETIREMENT_AGE - CURRENT_AGE
)


# ------------------------------------------------------------
# Existing EUR 50,000
#
# It grows for 15 years at the guaranteed 1%.
# ------------------------------------------------------------

FV_INITIAL_PER_PERSON = (

    INITIAL_CAPITAL_PER_PERSON
    *
    (1 + GUARANTEED_ACCUMULATION_RATE)
    ** YEARS_TO_RETIREMENT

)


# ------------------------------------------------------------
# Future contributions
#
# Contributions occur at the END of Years 1-10.
#
# Therefore:
#
# Year 1 contribution grows for 14 years
# Year 2 contribution grows for 13 years
# ...
# Year 10 contribution grows for 5 years
# ------------------------------------------------------------

FV_CONTRIBUTIONS_PER_PERSON = 0

contribution_details = []


for contribution_year in range(
    1,
    CONTRIBUTION_YEARS + 1
):

    years_of_growth = (

        YEARS_TO_RETIREMENT
        -
        contribution_year

    )

    future_value = (

        ANNUAL_CONTRIBUTION
        *
        (
            1
            +
            GUARANTEED_ACCUMULATION_RATE
        )
        ** years_of_growth

    )

    FV_CONTRIBUTIONS_PER_PERSON += future_value

    contribution_details.append({

        "Contribution_Year":
            contribution_year,

        "Contribution":
            ANNUAL_CONTRIBUTION,

        "Years_of_Growth":
            years_of_growth,

        "Value_at_Age_65":
            future_value

    })


# ------------------------------------------------------------
# Guaranteed capital per person
# ------------------------------------------------------------

GUARANTEED_CAPITAL_PER_PERSON = (

    FV_INITIAL_PER_PERSON
    +
    FV_CONTRIBUTIONS_PER_PERSON

)


# ------------------------------------------------------------
# Guaranteed capital for entire pool
# ------------------------------------------------------------

TOTAL_GUARANTEED_CAPITAL_AT_65 = (

    GUARANTEED_CAPITAL_PER_PERSON
    *
    TOTAL_POLICYHOLDERS

)


# ============================================================
# 5. LOAD GERMAN MORTALITY DATA
# ============================================================

mortality_input = pd.read_csv(
    MORTALITY_FILE
)


mortality_input = mortality_input[

    (
        mortality_input["Age"]
        >= RETIREMENT_AGE
    )

    &

    (
        mortality_input["Age"]
        < MAX_AGE
    )

].copy()


# ------------------------------------------------------------
# Check missing ages
# ------------------------------------------------------------

required_ages = set(
    range(
        RETIREMENT_AGE,
        MAX_AGE
    )
)

available_ages = set(
    mortality_input["Age"]
)

missing_ages = (
    required_ages
    -
    available_ages
)


if missing_ages:

    raise ValueError(

        f"Missing mortality rates for ages: "
        f"{sorted(missing_ages)}"

    )


# ============================================================
# 6. CREATE MORTALITY DICTIONARIES
# ============================================================

male_qx = dict(

    zip(

        mortality_input["Age"],

        mortality_input["Male_qx"]

    )

)


female_qx = dict(

    zip(

        mortality_input["Age"],

        mortality_input["Female_qx"]

    )

)


# ============================================================
# 7. SURVIVAL FUNCTION
# ============================================================

def calculate_survival(
    start_population,
    qx_by_age
):

    alive = float(
        start_population
    )

    rows = []


    for age in range(
        RETIREMENT_AGE,
        MAX_AGE
    ):

        qx = qx_by_age[age]

        expected_deaths = (
            alive * qx
        )

        alive_next_year = (
            alive
            -
            expected_deaths
        )


        rows.append({

            "Age":
                age,

            "Alive_Start":
                alive,

            "qx":
                qx,

            "Expected_Deaths":
                expected_deaths,

            "Alive_Next":
                alive_next_year

        })


        alive = (
            alive_next_year
        )


    return pd.DataFrame(
        rows
    )


# ============================================================
# 8. CALCULATE MALE/FEMALE SURVIVAL
# ============================================================

male = calculate_survival(
    N_MALE,
    male_qx
)

female = calculate_survival(
    N_FEMALE,
    female_qx
)


# ============================================================
# 9. BUILD PENSION MODEL
# ============================================================

model = pd.DataFrame()


model["Age"] = (
    male["Age"]
)


# ------------------------------------------------------------
# Pension year
#
# Year 1 = first year after retirement
# ------------------------------------------------------------

model["Pension_Year"] = (

    model["Age"]
    -
    RETIREMENT_AGE
    +
    1

)


model["Male_Alive"] = (
    male["Alive_Start"]
)

model["Female_Alive"] = (
    female["Alive_Start"]
)


model["Male_Deaths"] = (
    male["Expected_Deaths"]
)

model["Female_Deaths"] = (
    female["Expected_Deaths"]
)


model["Total_Alive"] = (

    model["Male_Alive"]
    +
    model["Female_Alive"]

)


model["Total_Deaths"] = (

    model["Male_Deaths"]
    +
    model["Female_Deaths"]

)


# ============================================================
# 10. NUMBER OF PENSION PAYMENTS
# ============================================================
#
# ASSUMPTION:
#
# First 10 pension years are guaranteed.
#
# Therefore:
#
# Pension Years 1-10:
# 100,000 pension payments.
#
# After Year 10:
# only surviving pensioners receive payments.
# ============================================================

def calculate_pension_units(row):

    if (
        row["Pension_Year"]
        <=
        PENSION_GUARANTEE_YEARS
    ):

        return (
            TOTAL_POLICYHOLDERS
        )

    else:

        return (
            row["Total_Alive"]
        )


model["Pension_Units"] = (

    model.apply(
        calculate_pension_units,
        axis=1
    )

)


# ============================================================
# 11. CALCULATE LEVEL PENSION AT AGE 65
# ============================================================
#
# We need an annual pension amount.
#
# For now:
#
# PV at age 65 of expected pension payments
# using the 1% conversion assumption
#
# =
#
# guaranteed accumulated capital at age 65.
# ============================================================

model["Conversion_Discount_Factor"] = (

    1
    /
    (
        1
        +
        PENSION_CONVERSION_RATE
    )
    ** model["Pension_Year"]

)


model["PV_1EUR_Pension_at_65"] = (

    model["Pension_Units"]
    *
    model[
        "Conversion_Discount_Factor"
    ]

)


POOL_ANNUITY_FACTOR = (

    model[
        "PV_1EUR_Pension_at_65"
    ].sum()

)


ANNUAL_PENSION_PER_PERSON = (

    TOTAL_GUARANTEED_CAPITAL_AT_65
    /
    POOL_ANNUITY_FACTOR

)


MONTHLY_PENSION_PER_PERSON = (

    ANNUAL_PENSION_PER_PERSON
    /
    12

)


# ============================================================
# 12. EXPECTED PENSION CASH FLOWS
# ============================================================

model[
    "Expected_Annual_Pension_Cashflow"
] = (

    model["Pension_Units"]
    *
    ANNUAL_PENSION_PER_PERSON

)


# ============================================================
# 13. READ EIOPA RISK-FREE CURVE DIRECTLY FROM ZIP
# ============================================================

with zipfile.ZipFile(
    EIOPA_ZIP_FILE,
    "r"
) as z:

    excel_bytes = z.read(
        EIOPA_EXCEL_FILE
    )


workbook = openpyxl.load_workbook(

    io.BytesIO(
        excel_bytes
    ),

    data_only=True

)


sheet = workbook[
    EIOPA_SHEET
]


# ============================================================
# 14. EXTRACT EURO SPOT CURVE
# ============================================================
#
# In the EIOPA file:
#
# Column B = maturity
# Column C = Euro spot rate
#
# Data begins at row 11.
# ============================================================

eiopa_curve = {}


row = 11


while True:

    maturity = (
        sheet.cell(
            row=row,
            column=2
        ).value
    )

    euro_rate = (
        sheet.cell(
            row=row,
            column=3
        ).value
    )


    if maturity is None:

        break


    if euro_rate is not None:

        eiopa_curve[
            int(maturity)
        ] = float(
            euro_rate
        )


    row += 1


# ============================================================
# 15. DISPLAY EIOPA CURVE INFORMATION
# ============================================================

EIOPA_LLP = (
    sheet["C5"].value
)

EIOPA_CONVERGENCE = (
    sheet["C6"].value
)

EIOPA_UFR = (
    sheet["C7"].value
)


# ============================================================
# 16. MAP PENSION CASH FLOWS TO TODAY
# ============================================================
#
# VERY IMPORTANT:
#
# Policyholders are 50 today.
#
# Pension begins at age 65.
#
# Therefore pension payments are NOT Year 1 from today.
#
# They begin roughly 15 years from now.
#
# Since we model the first annual payment at the END of the
# first pension year:
#
# Pension Year 1 = Year 16 from today
#
# Pension Year 2 = Year 17 from today
#
# etc.
#
# If you instead want the first payment exactly at age 65,
# change the line below to:
#
# model["Years_From_Today"] = 15 + model["Pension_Year"] - 1
#
# ============================================================

model["Years_From_Today"] = (

    YEARS_TO_RETIREMENT
    +
    model["Pension_Year"]

)


# ============================================================
# 17. GET EIOPA RATE FOR EACH LIABILITY CASH FLOW
# ============================================================

def get_eiopa_rate(year):

    year = int(year)

    if year not in eiopa_curve:

        raise ValueError(

            f"EIOPA curve does not contain "
            f"a rate for maturity {year} years."

        )

    return (
        eiopa_curve[year]
    )


model["EIOPA_Spot_Rate"] = (

    model[
        "Years_From_Today"
    ].apply(
        get_eiopa_rate
    )

)


# ============================================================
# 18. CALCULATE EIOPA DISCOUNT FACTORS
# ============================================================
#
# EIOPA file gives annual spot rates.
#
# Discount factor:
#
# DF(t) = 1 / (1 + r_t)^t
# ============================================================

model["EIOPA_Discount_Factor"] = (

    1
    /
    (
        1
        +
        model["EIOPA_Spot_Rate"]
    )
    **
    model["Years_From_Today"]

)


# ============================================================
# 19. CALCULATE PV TODAY OF EACH PENSION CASH FLOW
# ============================================================

model[
    "PV_Today_EIOPA"
] = (

    model[
        "Expected_Annual_Pension_Cashflow"
    ]

    *

    model[
        "EIOPA_Discount_Factor"
    ]

)


# ============================================================
# 20. TOTAL PV TODAY OF PENSION LIABILITY
# ============================================================

PV_PENSION_LIABILITY_TODAY = (

    model[
        "PV_Today_EIOPA"
    ].sum()

)


# ============================================================
# 21. PV OF PENSION LIABILITY AT AGE 65
# ============================================================
#
# This is based on our pension conversion assumption.
#
# It should equal the guaranteed accumulated capital at 65
# because that is how we solved for the pension.
# ============================================================

model[
    "PV_at_65_Conversion_Rate"
] = (

    model[
        "Expected_Annual_Pension_Cashflow"
    ]

    *

    model[
        "Conversion_Discount_Factor"
    ]

)


PV_PENSION_AT_65 = (

    model[
        "PV_at_65_Conversion_Rate"
    ].sum()

)


# ============================================================
# 22. TOTAL NOMINAL EXPECTED PAYMENTS
# ============================================================

TOTAL_NOMINAL_PENSION_PAYMENTS = (

    model[
        "Expected_Annual_Pension_Cashflow"
    ].sum()

)


# ============================================================
# 23. PRINT RESULTS
# ============================================================

print(
    "\n"
    +
    "=" * 75
)

print(
    "ALL-KANNS LIFE INSURANCE - "
    "EIOPA PENSION LIABILITY MODEL"
)

print(
    "=" * 75
)


# ------------------------------------------------------------
# Accumulation
# ------------------------------------------------------------

print(
    "\nACCUMULATION TO AGE 65"
)

print(
    "-" * 75
)


print(
    f"Initial capital per person: "
    f"EUR {INITIAL_CAPITAL_PER_PERSON:,.2f}"
)


print(
    f"Guaranteed return: "
    f"{GUARANTEED_ACCUMULATION_RATE * 100:.2f}%"
)


print(
    f"FV of initial EUR 50,000 at age 65: "
    f"EUR {FV_INITIAL_PER_PERSON:,.2f}"
)


print(
    f"FV of future contributions at age 65: "
    f"EUR {FV_CONTRIBUTIONS_PER_PERSON:,.2f}"
)


print(
    f"\nGuaranteed capital per person at 65: "
    f"EUR {GUARANTEED_CAPITAL_PER_PERSON:,.2f}"
)


print(
    f"Total guaranteed capital at 65: "
    f"EUR {TOTAL_GUARANTEED_CAPITAL_AT_65:,.2f}"
)


# ------------------------------------------------------------
# Pension
# ------------------------------------------------------------

print(
    "\nPENSION"
)

print(
    "-" * 75
)


print(
    f"Pension conversion rate assumption: "
    f"{PENSION_CONVERSION_RATE * 100:.2f}%"
)


print(
    f"Annual pension per person: "
    f"EUR {ANNUAL_PENSION_PER_PERSON:,.2f}"
)


print(
    f"Monthly pension per person: "
    f"EUR {MONTHLY_PENSION_PER_PERSON:,.2f}"
)


print(
    f"Total expected nominal lifetime payments: "
    f"EUR {TOTAL_NOMINAL_PENSION_PAYMENTS:,.2f}"
)


# ------------------------------------------------------------
# EIOPA
# ------------------------------------------------------------

print(
    "\nEIOPA CURVE"
)

print(
    "-" * 75
)


print(
    "Curve: EUR RFR spot without VA"
)


print(
    f"LLP: {EIOPA_LLP} years"
)


print(
    f"Convergence: {EIOPA_CONVERGENCE} years"
)


print(
    f"UFR: {EIOPA_UFR:.2f}%"
)


print(
    f"First EIOPA rate used: "
    f"{model.iloc[0]['EIOPA_Spot_Rate'] * 100:.4f}%"
)


print(
    f"First liability maturity from today: "
    f"{int(model.iloc[0]['Years_From_Today'])} years"
)


# ------------------------------------------------------------
# Liability valuation
# ------------------------------------------------------------

print(
    "\nLIABILITY VALUATION"
)

print(
    "-" * 75
)


print(
    f"PV at age 65 under pension conversion assumption: "
    f"EUR {PV_PENSION_AT_65:,.2f}"
)


print(
    f"PV TODAY using EIOPA EUR spot curve: "
    f"EUR {PV_PENSION_LIABILITY_TODAY:,.2f}"
)


print(
    f"\nCurrent assets: "
    f"EUR {INITIAL_CAPITAL_PER_PERSON * TOTAL_POLICYHOLDERS:,.2f}"
)


# ============================================================
# 24. PRINT CASH-FLOW TABLE
# ============================================================

output_columns = [

    "Pension_Year",

    "Age",

    "Years_From_Today",

    "Male_Alive",

    "Female_Alive",

    "Total_Alive",

    "Total_Deaths",

    "Pension_Units",

    "Expected_Annual_Pension_Cashflow",

    "EIOPA_Spot_Rate",

    "EIOPA_Discount_Factor",

    "PV_Today_EIOPA"

]


print(
    "\n"
    +
    "=" * 75
)

print(
    "EXPECTED PENSION CASH FLOWS AND EIOPA PRESENT VALUES"
)

print(
    "=" * 75
    +
    "\n"
)


display_model = (
    model[output_columns].copy()
)


# Convert EIOPA rate to percentage for easier reading

display_model[
    "EIOPA_Spot_Rate"
] = (

    display_model[
        "EIOPA_Spot_Rate"
    ]
    *
    100

)


print(

    display_model

    .round(4)

    .to_string(
        index=False
    )

)


# ============================================================
# 25. EXPORT TO EXCEL
# ============================================================

model[
    output_columns
].to_excel(

    "pension_liability_EIOPA.xlsx",

    index=False

)


pd.DataFrame(
    contribution_details
).to_excel(

    "contribution_accumulation.xlsx",

    index=False

)


# Export EIOPA curve as well

eiopa_export = pd.DataFrame({

    "Maturity":
        list(
            eiopa_curve.keys()
        ),

    "EUR_Spot_Rate":
        list(
            eiopa_curve.values()
        )

})


eiopa_export.to_excel(

    "EIOPA_EUR_curve.xlsx",

    index=False

)


print(
    "\nFiles created:"
)

print(
    "1. pension_liability_EIOPA.xlsx"
)

print(
    "2. contribution_accumulation.xlsx"
)

print(
    "3. EIOPA_EUR_curve.xlsx"
)